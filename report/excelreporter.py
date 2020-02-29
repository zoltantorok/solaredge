'''
Created on 22.11.2019

@author: Zoli
'''
import openpyxl
from os import path
from pathlib import Path
import atexit

class ExcelReporter(object):
    '''
    classdocs
    '''


    def __init__(self, outputFilename):
        '''
        Constructor
        '''
        self.outputFilename = outputFilename
        name, ext = path.splitext(self.outputFilename)
        
        excelExt = '.xlsx'
        if ext.casefold() != excelExt:
            self.outputFilename = path.join(Path(self.outputFilename).parent, name) + excelExt 
        
        self.workBook = openpyxl.Workbook()
        atexit.register(self.storeWorkbook)
        
    def storeWorkbook(self):
        self.workBook.save(self.outputFilename)

    def writeEnergyData(self, energyData, energyTypes, costCalculator):
        
        lineNumber = 1
        lastYear = 0
        lastWeekNumber = -1
        firstTimestampOfWeek = None
        
        timestamps = list(energyData.keys())
        for timestamp in timestamps:
            if lastYear == 0:
                lastYear = timestamp.year
                worksheet = self.workBook.active
                worksheet.title = str(timestamp.year)
            elif timestamp.year != lastYear:
                self.writeYearTotals(worksheet, lastYear, energyData, energyTypes, costCalculator)
                lastYear = timestamp.year
                worksheet = self.workBook.create_sheet(str(timestamp.year))
                lineNumber = 1
                lastWeekNumber = -1
                firstTimestampOfWeek = None
            
            if timestamp.date().isocalendar()[1] != lastWeekNumber:
                # New week
                # Calculate total
                if firstTimestampOfWeek is not None:
                    worksheet['I' + str(lineNumber)] = 'Total (KWh)'
                    # Add total week cost/refund
                    weekEnergyData = self.getDictSubsetByKeyRange(energyData, timestamps[timestamps.index(firstTimestampOfWeek):timestamps.index(timestamp)])
                    self.writeWeekTotals(worksheet, lineNumber, weekEnergyData, energyTypes, costCalculator)
                        
                lastWeekNumber = timestamp.date().isocalendar()[1]
                lineNumber = lineNumber + 10
                firstTimestampOfWeek = timestamp
            
            value = energyData[timestamp]
            
            column = self.getColumnForWeekday(timestamp.weekday())
            for val in range(len(value)):
                worksheet[column + str(lineNumber)] = timestamp.strftime("%A")
                worksheet[column + str(lineNumber + 1)] = timestamp.day
                worksheet['A' + str(lineNumber + 1)] = timestamp.strftime("%B") 
                worksheet['A' + str(lineNumber + val + 2)] = energyTypes[val]
                cell = column + str(lineNumber + val + 2)
                if worksheet[cell].value is None:
                    worksheet[cell] = value[val] / 1000.0
                else:
                    worksheet[cell] = worksheet[cell].value + value[val] / 1000.0

        if firstTimestampOfWeek != timestamp:
            worksheet['I' + str(lineNumber)] = 'Total (KWh)'
            # Add total week cost/refund
            weekEnergyData = self.getDictSubsetByKeyRange(energyData, timestamps[timestamps.index(firstTimestampOfWeek):timestamps.index(timestamp)])
            self.writeWeekTotals(worksheet, lineNumber, weekEnergyData, energyTypes, costCalculator)
        
        if lastYear == timestamp.year:
            self.writeYearTotals(worksheet, timestamp.year, energyData, energyTypes, costCalculator)
                
    def writeWeekTotals(self, worksheet, lineNumber, weekEnergyData, energyTypes, costCalculator):
        worksheet['I' + str(lineNumber)] = 'Total (KWh)'
        # Add total week cost/refund
        weekCost = costCalculator.calculateCost(weekEnergyData, energyTypes)
        worksheet['K' + str(lineNumber)] = 'Week cost'
        worksheet['L' + str(lineNumber)] = weekCost
        worksheet['M' + str(lineNumber)] = 'currency'
        sumEnergyWeek = self.sumEnergyData(weekEnergyData)
        self.writeSelfSufficiency(worksheet, sumEnergyWeek, energyTypes, startingCell='K', lineNumber=lineNumber + 1)
        self.writeOwnConsumption(worksheet, sumEnergyWeek, energyTypes, startingCell='K', lineNumber=lineNumber + 2)
        for val in range(len(sumEnergyWeek)):
            worksheet['I' + str(lineNumber + val + 2)] = sumEnergyWeek[val] / 1000.0
        
    def writeSelfSufficiency(self, worksheet, energyData, energyTypes, startingCell, lineNumber):
        worksheet[startingCell + str(lineNumber)] = 'Self-sufficiency'
        worksheet[chr(ord(startingCell) + 1) + str(lineNumber)] = energyData[energyTypes.index('SelfConsumption')] / energyData[energyTypes.index('Consumption')] * 100
        worksheet[chr(ord(startingCell) + 2) + str(lineNumber)] = '%'

    def writeOwnConsumption(self, worksheet, energyData, energyTypes, startingCell, lineNumber):
        worksheet[startingCell + str(lineNumber)] = 'Own-consumption'
        valueCell = chr(ord(startingCell) + 1) + str(lineNumber)
        production = energyData[energyTypes.index('Production')]
        if production > 0:
            feedIn = energyData[energyTypes.index('FeedIn')]
            worksheet[valueCell] = 100.0 * (production - feedIn) / production
        else:
            worksheet[valueCell] = 0
        worksheet[chr(ord(startingCell) + 2) + str(lineNumber)] = '%'
    
    def writeYearTotals(self, worksheet, year, energyData, energyTypes, costCalculator):
        yearTimestampBorders = []
        timestamps = list(energyData.keys())
        
        for timestamp in timestamps:
            if timestamp.year == year:
                if len(yearTimestampBorders) == 0:
                    yearTimestampBorders.append(timestamp)
                    yearTimestampBorders.append(timestamp)
                    continue
                yearTimestampBorders[1] = timestamp
        
        yearEnergyData = self.getDictSubsetByKeyRange(energyData, timestamps[timestamps.index(yearTimestampBorders[0]):timestamps.index(yearTimestampBorders[1])])
        yearEnergyCost = costCalculator.calculateCost(yearEnergyData, energyTypes)
        worksheet['A3'] = 'Year totals'
        worksheet['B2'] = 'Start date'
        worksheet['C2'] = 'End date'
        worksheet['B3'] = yearTimestampBorders[0].date().strftime('%x')
        worksheet['C3'] = yearTimestampBorders[1].date().strftime('%x')
        worksheet['A5'] = 'Cost'
        worksheet['B5'] = yearEnergyCost
        worksheet['C5'] = 'currency'
        yearEnergySum = self.sumEnergyData(yearEnergyData)
        worksheet['A6'] = 'Purchased (KWh)'
        worksheet['B6'] = yearEnergySum[energyTypes.index('Purchased')] / 1000.0
        worksheet['A7'] = 'FeedIn (KWh)'
        worksheet['B7'] = yearEnergySum[energyTypes.index('FeedIn')] / 1000.0
        self.writeSelfSufficiency(worksheet, yearEnergySum, energyTypes, startingCell='A', lineNumber=8)
        self.writeOwnConsumption(worksheet, yearEnergySum, energyTypes, startingCell='A', lineNumber=9)

    def getDictSubsetByKeyRange(self, dictionary, keyRange):
        outDict = {}
        for key in dictionary.keys():
            if key in keyRange:
                outDict[key] = dictionary[key]
        return outDict
    
    def getColumnForWeekday(self, weekday):
        return chr(ord('B') + weekday)
    
    def sumEnergyData(self, energyData):
        keys = list(energyData.keys())
        sumValues = [0] * len(energyData[keys[0]])
        
        for key in keys:
            value = energyData[key]
            for idx in range(len(value)):
                sumValues[idx] = sumValues[idx] + value[idx]
        
        return sumValues
    























